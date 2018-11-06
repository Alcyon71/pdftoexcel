# -*- coding: utf-8 -*-
import tabula
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
#TODO : Choix du fichier à lire

# Read pdf into DataFrame
df = tabula.read_pdf("test.pdf", pages='all', multiple_tables=True)
#on supprime les 3 premieres lignes
df = [df[i].iloc[3:,] for i in range(len(df))]

#On merge les tableaux de df
merge = pd.DataFrame()
merge = pd.concat([i for i in df],ignore_index=True)

# on va essayer de filtrer les identifiants des échantillons, soit bon, soit coulée soit fils de la ligne précédente
for row in merge.itertuples():
    print(row)
    #print(merge.iloc[row[0],3])
    #Un casier a maximum 4 chiffre
    if len(row[4]) > 4:
        merge.iloc[row[0],3] = merge.iloc[row[0],3].replace(" ", "-")
    else:
        merge.iloc[row[0], 3] = merge.iloc[row[0]-1, 3]

#TODO : Si la date n'est pas en colonne 9(10 dans le namedtuple), la copier de la colonne 10(11 dans le namedtuple), puis sup la colonne 10
    print(type(row[10]))
    if type(row[10]) is float:
        merge.iloc[row[0], 9] = merge.iloc[row[0], 10]

#On supprime la colonne 10 et 4
merge = merge.drop([4, 10], axis=1)
print(merge)

#Ajout d'une feuille et copie des valeurs
wb = load_workbook(filename='merge.xlsm', keep_vba=True)
ws = wb.create_sheet('merge')

for r in dataframe_to_rows(merge, index=False, header=False):
    ws.append(r)

wb.save('merge.xlsm')

# writer = pd.ExcelWriter('merge.xlsx', )
# merge.to_excel(writer, 'merge')
# writer.save()




#writer = pd.ExcelWriter('output.xlsx')
# i = 0
# for tab in df:
#     i += 1
#     tab = tab.iloc[3:,]
#     tab.to_excel(writer, 'Sheet'+str(i))

#writer.save()


